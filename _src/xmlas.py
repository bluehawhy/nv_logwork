#!/usr/bin/python
# excel.py
# -*- coding: utf-8 -*-

'''
Created on 2019. 1. 14.
@author: miskang
#update list
2026-07-13 : openpyxl 최신 API 반영 및 self. 안티패턴 수정
'''

import openpyxl


class Workbook(object):
    def __init__(self, file, read_only=False, data_only=False):
        # 파라미터 기본값을 주어 호출 시 유연성을 높였습니다.
        self.wb = openpyxl.load_workbook(file, read_only=read_only, data_only=data_only)
    
    def create_new_sheet(self, sheet_name):
        # self.ws 대신 지역 변수로 시트를 생성해 반환합니다.
        ws = self.wb.create_sheet(sheet_name)
        return ws

    def remove_sheet(self, sheet_name):
        # 구형 get_sheet_by_name 대신 대괄호 키 접근([sheet_name])을 사용합니다.
        if sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]
            self.wb.remove(ws)
        return 0

    def get_sheet_list(self):
        return self.wb.sheetnames

    def get_first_row(self, sheet_name):
        ws = self.wb[sheet_name]
        # 리스트 컴프리헨션(List Comprehension)을 사용해 코드를 간결하게 축약했습니다.
        first_row_values = [cell.value for cell in ws[1]]
        return first_row_values

    def get_worksheet(self, sheet_name):
        return self.wb[sheet_name]

    def change_cell_data(self, ws, col, row, val):
        # col이 문자(예: 'A')로 들어오는지, 숫자(예: 1)로 들어오는지에 따라 
        # openpyxl이 유연하게 처리할 수 있도록 cell() 메서드를 유지합니다.
        if isinstance(col, str):
            ws[f"{col}{row}"] = val
        else:
            ws.cell(row=row, column=col, value=val)
        return 0

    def save_workbook(self, file):
        try:
            self.wb.save(file)
        except PermissionError:
            # f-string을 사용해 문자열 포매팅을 깔끔하게 정리했습니다.
            alternative_file = str(file).replace('.xlsx', '_temp.xlsx')
            self.wb.save(alternative_file)
        return 0

    def close_workbook(self):
        self.wb.close()
        return 0